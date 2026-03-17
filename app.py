import random
import string
from datetime import datetime
import csv
import io
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, flash, g, Response
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import NullPool
from dotenv import load_dotenv
import os
from analytics.predictor import calcular_metricas_reposicao

load_dotenv()

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "sua_chave_secreta_aqui")

if os.getenv("VERCEL") or os.getenv("FLASK_ENV") == "production":
    app.config.update(
        SESSION_COOKIE_SECURE=True,
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
    )

# Configuração do banco de dados Neon
DATABASE_URL = os.getenv("DATABASE_URL")
engine = None
SessionLocal = None
DB_BOOT_ERROR = None

try:
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL não configurada. Defina a variável de ambiente na Vercel.")

    # Para ambientes serverless como Vercel, use NullPool
    if os.getenv("VERCEL"):
        engine = create_engine(DATABASE_URL, poolclass=NullPool, future=True)
    else:
        engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_size=5, max_overflow=10, future=True)

    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
except Exception as db_init_error:
    DB_BOOT_ERROR = str(db_init_error)

@app.before_request
def open_session():
    """Abre sessão do banco antes de cada request"""
    if SessionLocal is None:
        return jsonify({'error': 'Configuração de banco inválida no servidor.', 'detail': DB_BOOT_ERROR}), 500
    g.db = SessionLocal()

@app.teardown_request
def close_session(exception=None):
    """Fecha sessão do banco após cada request"""
    db = getattr(g, 'db', None)
    if db is not None:
        try:
            if exception:
                db.rollback()
            db.close()
        except Exception:
            pass

def login_required(f):
    def wrapper(*args, **kwargs):
        if 'user' not in session:
            flash('Você precisa estar logado para acessar essa página.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

def admin_required(f):
    def wrapper(*args, **kwargs):
        if 'user' not in session or session['user']['role'] != 'admin':
            flash('Acesso restrito a administradores.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        try:
            row = g.db.execute(text("""
                SELECT * FROM usuarios WHERE username = :username AND password = :password
            """), {"username": username, "password": password}).mappings().first()
            
            if row:
                user_data = dict(row)
                if user_data.get('must_change_password'):
                    session['user_temp'] = user_data
                    flash('Você deve trocar sua senha antes de continuar.', 'warning')
                    return redirect(url_for('trocar_senha'))
                session['user'] = user_data
                flash('Login realizado com sucesso!', 'success')
                if user_data['role'] == 'admin':
                    return redirect(url_for('usuarios'))
                return redirect(url_for('empresa'))
            flash('Usuário ou senha incorretos.', 'error')
        except Exception as e:
            flash('Erro ao fazer login: {}'.format(str(e)), 'error')
    
    return render_template('login.html')

@app.route('/trocar_senha', methods=['GET', 'POST'])
def trocar_senha():
    if 'user_temp' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        nova = request.form['nova_senha']
        confirmar = request.form['confirmar_senha']
        
        if nova != confirmar:
            flash('As senhas não coincidem.', 'error')
            return render_template('trocar_senha.html')
        
        try:
            user_id = session['user_temp']['id']
            g.db.execute(text("""
                UPDATE usuarios SET password = :password, must_change_password = FALSE 
                WHERE id = :id
            """), {"password": nova, "id": user_id})
            g.db.commit()
            
            # Busca o usuário atualizado
            row = g.db.execute(text("""
                SELECT * FROM usuarios WHERE id = :id
            """), {"id": user_id}).mappings().first()
            
            session.pop('user_temp')
            session['user'] = dict(row)
            flash('Senha alterada com sucesso! Faça login novamente.', 'success')
            return redirect(url_for('logout'))
        except Exception as e:
            flash('Erro ao alterar senha: {}'.format(str(e)), 'error')
    
    return render_template('trocar_senha.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    flash('Você foi deslogado.', 'success')
    return redirect(url_for('login'))

@app.route('/usuarios', methods=['GET', 'POST'])
@admin_required
def usuarios():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        try:
            # Verifica se usuário já existe
            existing = g.db.execute(text("""
                SELECT id FROM usuarios WHERE username = :username
            """), {"username": username}).mappings().first()
            
            if existing:
                flash('Nome de usuário já existe.', 'error')
            else:
                g.db.execute(text("""
                    INSERT INTO usuarios (username, password, role) 
                    VALUES (:username, :password, 'user')
                """), {"username": username, "password": password})
                g.db.commit()
                flash('Usuário cadastrado com sucesso!', 'success')
        except Exception as e:
            flash('Erro ao cadastrar usuário: {}'.format(str(e)), 'error')
        
        return redirect(url_for('usuarios'))
    
    try:
        rows = g.db.execute(text("""
            SELECT * FROM usuarios ORDER BY id ASC
        """)).mappings().all()
        usuarios_data = [dict(row) for row in rows]
    except Exception as e:
        flash('Erro ao carregar usuários: {}'.format(str(e)), 'error')
        usuarios_data = []
    
    return render_template('usuarios.html', usuarios=usuarios_data)

@app.route('/usuarios/update/<int:id>', methods=['POST'])
@admin_required
def update_usuario(id):
    try:
        new_password = request.form['password']
        g.db.execute(text("""
            UPDATE usuarios SET password = :password WHERE id = :id
        """), {"password": new_password, "id": id})
        g.db.commit()
        return jsonify({'success': 'Senha atualizada com sucesso!'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/usuarios/delete/<int:id>', methods=['POST'])
@admin_required
def delete_usuario(id):
    try:
        # Busca dados do usuário
        row = g.db.execute(text("""
            SELECT username, password, must_change_password FROM usuarios WHERE id = :id
        """), {"id": id}).mappings().first()
        
        if not row:
            return jsonify({'error': 'Usuário não encontrado.'}), 404
        
        user = dict(row)
        
        if user.get('username') == 'admin':
            return jsonify({'error': 'Não é possível excluir o usuário admin.'}), 400
        
        if user.get('must_change_password'):
            return jsonify({'error': "Não é possível excluir usuário enquanto utiliza uma senha temporária."}), 400
        
        current_password = request.form.get('current_password') or request.args.get('current_password')
        if not current_password:
            return jsonify({'error': 'É necessário informar a senha atual do usuário para exclusão.'}), 400
        
        stored_pw = user.get('password') or ''
        if str(current_password) != str(stored_pw):
            return jsonify({'error': 'Senha incorreta. Exclusão não autorizada.'}), 403

        g.db.execute(text("DELETE FROM usuarios WHERE id = :id"), {"id": id})
        g.db.commit()
        return jsonify({'success': 'Usuário excluído com sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/usuarios/reset/<int:id>', methods=['POST'])
@admin_required
def reset_usuario(id):
    try:
        temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        
        try:
            g.db.execute(text("""
                UPDATE usuarios SET password = :password, must_change_password = TRUE 
                WHERE id = :id
            """), {"password": temp_password, "id": id})
            g.db.commit()
            return jsonify({'temp_password': temp_password})
        except Exception as e:
            # Fallback se coluna must_change_password não existir
            try:
                g.db.execute(text("""
                    UPDATE usuarios SET password = :password WHERE id = :id
                """), {"password": temp_password, "id": id})
                g.db.commit()
                return jsonify({
                    'temp_password': temp_password,
                    'warning': "coluna 'must_change_password' não encontrada; senha atualizada mas usuário NÃO foi marcado para trocar senha."
                })
            except Exception as e2:
                return jsonify({'error': 'Falha ao atualizar senha (fallback): {}'.format(str(e2))}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/', methods=['GET', 'POST'])
@login_required
def empresa():
    can_edit = session['user']['role'] == 'admin'
    
    try:
        row = g.db.execute(text("SELECT * FROM empresa LIMIT 1")).mappings().first()
        empresa_data = dict(row) if row else None
    except Exception as e:
        flash('Erro ao carregar dados da empresa: {}'.format(str(e)), 'error')
        empresa_data = None
    
    if request.method == 'POST':
        if not can_edit:
            flash('Apenas administradores podem alterar os dados da empresa.', 'error')
            return redirect(url_for('empresa'))
        
        try:
            data = {
                "cnpj": request.form['cnpj'],
                "razao_social": request.form['razao_social'],
                "nome_fantasia": request.form['nome_fantasia'],
                "rua": request.form['rua'],
                "numero": request.form['numero'],
                "complemento": request.form['complemento'],
                "cep": request.form['cep'],
                "bairro": request.form['bairro'],
                "cidade": request.form['cidade'],
                "estado": request.form['estado'],
                "telefone": request.form['telefone']
            }
            
            if empresa_data:
                g.db.execute(text("""
                    UPDATE empresa SET cnpj = :cnpj, razao_social = :razao_social, 
                    nome_fantasia = :nome_fantasia, rua = :rua, numero = :numero, 
                    complemento = :complemento, cep = :cep, bairro = :bairro, 
                    cidade = :cidade, estado = :estado, telefone = :telefone 
                    WHERE id = :id
                """), {**data, "id": empresa_data['id']})
            else:
                g.db.execute(text("""
                    INSERT INTO empresa (cnpj, razao_social, nome_fantasia, rua, numero, 
                    complemento, cep, bairro, cidade, estado, telefone) 
                    VALUES (:cnpj, :razao_social, :nome_fantasia, :rua, :numero, 
                    :complemento, :cep, :bairro, :cidade, :estado, :telefone)
                """), data)
            g.db.commit()
            flash('Dados da empresa salvos com sucesso!', 'success')
        except Exception as e:
            flash('Erro ao salvar dados da empresa: {}'.format(str(e)), 'error')
        
        return redirect(url_for('empresa'))
    
    return render_template('empresa.html', empresa=empresa_data, can_edit=can_edit)

@app.route('/fornecedores', methods=['GET', 'POST'])
@login_required
def fornecedores():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))
    
    if request.method == 'POST':
        try:
            data = {
                "cnpj": request.form['cnpj'],
                "razao_social": request.form['razao_social'],
                "nome_fantasia": request.form['nome_fantasia'],
                "rua": request.form['rua'],
                "numero": request.form['numero'],
                "complemento": request.form['complemento'],
                "cep": request.form['cep'],
                "bairro": request.form['bairro'],
                "cidade": request.form['cidade'],
                "estado": request.form['estado'],
                "telefone": request.form['telefone'],
                "representante": request.form['representante']
            }
            
            if 'id' in request.form and request.form['id']:
                g.db.execute(text("""
                    UPDATE fornecedor SET cnpj = :cnpj, razao_social = :razao_social, 
                    nome_fantasia = :nome_fantasia, rua = :rua, numero = :numero, 
                    complemento = :complemento, cep = :cep, bairro = :bairro, 
                    cidade = :cidade, estado = :estado, telefone = :telefone, 
                    representante = :representante WHERE id = :id
                """), {**data, "id": request.form['id']})
            else:
                g.db.execute(text("""
                    INSERT INTO fornecedor (cnpj, razao_social, nome_fantasia, rua, numero, 
                    complemento, cep, bairro, cidade, estado, telefone, representante) 
                    VALUES (:cnpj, :razao_social, :nome_fantasia, :rua, :numero, 
                    :complemento, :cep, :bairro, :cidade, :estado, :telefone, :representante)
                """), data)
            g.db.commit()
            flash('Fornecedor salvo com sucesso!', 'success')
        except Exception as e:
            flash('Erro ao salvar fornecedor: {}'.format(str(e)), 'error')
        
        return redirect(url_for('fornecedores'))
    
    filtro_status = request.args.get('status', 'ativos')
    try:
        if filtro_status == 'inativos':
            sql = "SELECT id, nome_fantasia, razao_social, telefone, representante, ativo FROM fornecedor WHERE ativo = FALSE ORDER BY id ASC"
        elif filtro_status == 'todos':
            sql = "SELECT id, nome_fantasia, razao_social, telefone, representante, ativo FROM fornecedor ORDER BY id ASC"
        else:
            filtro_status = 'ativos'
            sql = "SELECT id, nome_fantasia, razao_social, telefone, representante, ativo FROM fornecedor WHERE ativo = TRUE ORDER BY id ASC"
        rows = g.db.execute(text(sql)).mappings().all()
        fornecedores_data = [dict(row) for row in rows]
    except Exception as e:
        flash('Erro ao carregar fornecedores: {}'.format(str(e)), 'error')
        fornecedores_data = []
    
    return render_template('fornecedores.html', fornecedores=fornecedores_data, filtro_status=filtro_status)

@app.route('/fornecedores/delete/<int:id>', methods=['POST'])
@login_required
def delete_fornecedor(id):
    try:
        # Verifica se existem movimentações vinculadas ao fornecedor
        mov_count = g.db.execute(text("""
            SELECT COUNT(*) as total FROM movimentacao 
            WHERE parceiro_id = :id AND parceiro_tipo = 'fornecedor'
        """), {"id": id}).mappings().first()
        
        if mov_count and int(mov_count['total']) > 0:
            return jsonify({
                'error': 'Não é possível excluir este fornecedor pois existem {} movimentação(ões) vinculada(s). Sugerimos inativá-lo em vez de excluir.'.format(int(mov_count['total'])),
                'has_movimentacoes': True,
                'total_movimentacoes': int(mov_count['total'])
            }), 400
        
        g.db.execute(text("DELETE FROM fornecedor WHERE id = :id"), {"id": id})
        g.db.commit()
        return jsonify({'success': 'Fornecedor excluído com sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/fornecedores/toggle/<int:id>', methods=['POST'])
@login_required
def toggle_fornecedor(id):
    try:
        row = g.db.execute(text("SELECT ativo FROM fornecedor WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Fornecedor não encontrado'}), 404
        
        novo_status = not row['ativo']
        g.db.execute(text("UPDATE fornecedor SET ativo = :ativo WHERE id = :id"), {"ativo": novo_status, "id": id})
        g.db.commit()
        status_text = 'ativado' if novo_status else 'inativado'
        return jsonify({'success': 'Fornecedor {} com sucesso'.format(status_text), 'ativo': novo_status})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/fornecedores/edit/<int:id>', methods=['GET'])
@login_required
def get_fornecedor(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        row = g.db.execute(text("SELECT * FROM fornecedor WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Fornecedor não encontrado'}), 404
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/clientes', methods=['GET', 'POST'])
@login_required
def clientes():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))
    
    if request.method == 'POST':
        try:
            data = {
                'nome': request.form['nome'],
                'documento': request.form['documento'],
                'rua': request.form.get('rua',''),
                'numero': request.form.get('numero',''),
                'complemento': request.form.get('complemento',''),
                'cep': request.form.get('cep',''),
                'bairro': request.form.get('bairro',''),
                'cidade': request.form.get('cidade',''),
                'estado': request.form.get('estado',''),
                'telefone': request.form.get('telefone','')
            }
            
            # Validação do documento
            raw_doc = request.form['documento']
            doc_digits = ''.join(filter(str.isdigit, raw_doc))
            if len(doc_digits) not in (11,14):
                flash('Documento inválido. Informe CPF (11 dígitos) ou CNPJ (14 dígitos).', 'error')
                return redirect(url_for('clientes'))
            
            data['documento'] = raw_doc
            
            if 'id' in request.form and request.form['id']:
                g.db.execute(text("""
                    UPDATE cliente SET nome = :nome, documento = :documento, rua = :rua, 
                    numero = :numero, complemento = :complemento, cep = :cep, bairro = :bairro, 
                    cidade = :cidade, estado = :estado, telefone = :telefone WHERE id = :id
                """), {**data, "id": request.form['id']})
            else:
                g.db.execute(text("""
                    INSERT INTO cliente (nome, documento, rua, numero, complemento, cep, 
                    bairro, cidade, estado, telefone) VALUES (:nome, :documento, :rua, 
                    :numero, :complemento, :cep, :bairro, :cidade, :estado, :telefone)
                """), data)
            g.db.commit()
            flash('Cliente salvo com sucesso!', 'success')
        except Exception as e:
            flash('Erro ao salvar cliente: {}'.format(str(e)), 'error')
        
        return redirect(url_for('clientes'))

    filtro_status = request.args.get('status', 'ativos')
    try:
        if filtro_status == 'inativos':
            sql = "SELECT id, nome, documento, telefone, ativo FROM cliente WHERE ativo = FALSE ORDER BY id ASC"
        elif filtro_status == 'todos':
            sql = "SELECT id, nome, documento, telefone, ativo FROM cliente ORDER BY id ASC"
        else:
            filtro_status = 'ativos'
            sql = "SELECT id, nome, documento, telefone, ativo FROM cliente WHERE ativo = TRUE ORDER BY id ASC"
        rows = g.db.execute(text(sql)).mappings().all()
        clientes_data = [dict(row) for row in rows]
    except Exception as e:
        flash('Erro ao carregar clientes: {}'.format(str(e)), 'error')
        clientes_data = []
    
    return render_template('clientes.html', clientes=clientes_data, filtro_status=filtro_status)

@app.route('/clientes/delete/<int:id>', methods=['POST'])
@login_required
def delete_cliente(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        # Verifica se existem movimentações vinculadas ao cliente
        mov_count = g.db.execute(text("""
            SELECT COUNT(*) as total FROM movimentacao 
            WHERE parceiro_id = :id AND parceiro_tipo = 'cliente'
        """), {"id": id}).mappings().first()
        
        if mov_count and int(mov_count['total']) > 0:
            return jsonify({
                'error': 'Não é possível excluir este cliente pois existem {} movimentação(ões) vinculada(s). Sugerimos inativá-lo em vez de excluir.'.format(int(mov_count['total'])),
                'has_movimentacoes': True,
                'total_movimentacoes': int(mov_count['total'])
            }), 400
        
        g.db.execute(text("DELETE FROM cliente WHERE id = :id"), {"id": id})
        g.db.commit()
        return jsonify({'success': 'Cliente excluído com sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/clientes/toggle/<int:id>', methods=['POST'])
@login_required
def toggle_cliente(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        row = g.db.execute(text("SELECT ativo FROM cliente WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Cliente não encontrado'}), 404
        
        novo_status = not row['ativo']
        g.db.execute(text("UPDATE cliente SET ativo = :ativo WHERE id = :id"), {"ativo": novo_status, "id": id})
        g.db.commit()
        status_text = 'ativado' if novo_status else 'inativado'
        return jsonify({'success': 'Cliente {} com sucesso'.format(status_text), 'ativo': novo_status})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/clientes/edit/<int:id>', methods=['GET'])
@login_required
def get_cliente(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        row = g.db.execute(text("SELECT * FROM cliente WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Cliente não encontrado'}), 404
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/produtos', methods=['GET', 'POST'])
@login_required
def produtos():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))
    
    if request.method == 'POST':
        try:
            descricao = request.form['descricao']
            estoque_minimo_raw = request.form.get('estoque_minimo')

            def _parse_planejamento(valor):
                if valor is None:
                    return None
                valor_limpo = str(valor).strip()
                if valor_limpo == '':
                    return None
                numero = int(valor_limpo)
                return max(0, numero)

            estoque_minimo = _parse_planejamento(estoque_minimo_raw)
            possui_colunas_planejamento = _produto_tem_colunas_planejamento()
            
            if 'id' in request.form and request.form['id']:
                # Atualizar produto existente
                prod_id = request.form['id']
                # Busca a descrição atual
                row = g.db.execute(text("SELECT descricao FROM produto WHERE id = :id"), {"id": prod_id}).mappings().first()
                if row:
                    old_desc = row['descricao']
                    # Atualiza todos os produtos com a mesma descrição
                    if possui_colunas_planejamento:
                        g.db.execute(text("""
                            UPDATE produto
                            SET descricao = :new_desc,
                                estoque_minimo = :estoque_minimo
                            WHERE descricao = :old_desc
                        """), {
                            "new_desc": descricao,
                            "old_desc": old_desc,
                            "estoque_minimo": estoque_minimo
                        })
                    else:
                        g.db.execute(text("""
                            UPDATE produto SET descricao = :new_desc WHERE descricao = :old_desc
                        """), {"new_desc": descricao, "old_desc": old_desc})
            else:
                # Inserir novo produto
                if possui_colunas_planejamento:
                    g.db.execute(text("""
                        INSERT INTO produto (descricao, quantidade, local, estoque_minimo)
                        VALUES (:descricao, 0, '', :estoque_minimo)
                    """), {
                        "descricao": descricao,
                        "estoque_minimo": estoque_minimo
                    })
                else:
                    g.db.execute(text("""
                        INSERT INTO produto (descricao, quantidade, local) VALUES (:descricao, 0, '')
                    """), {"descricao": descricao})
            
            g.db.commit()
            flash('Produto salvo com sucesso!', 'success')
        except Exception as e:
            flash('Erro ao salvar produto: {}'.format(str(e)), 'error')
        
        return redirect(url_for('produtos'))
    
    try:
        possui_colunas_planejamento = _produto_tem_colunas_planejamento()
        if possui_colunas_planejamento:
            rows = g.db.execute(text("""
                SELECT id, descricao, quantidade, local, estoque_minimo
                FROM produto ORDER BY id ASC
            """)).mappings().all()
        else:
            rows = g.db.execute(text("""
                SELECT id, descricao, quantidade, local
                FROM produto ORDER BY id ASC
            """)).mappings().all()
        
        # Agrupa produtos por descrição
        produtos_map = {}
        for row in rows:
            r = dict(row)
            key = r.get('descricao') or 'produto_{}'.format(r.get("id"))
            if key not in produtos_map:
                produtos_map[key] = {
                    'id': r.get('id'),
                    'descricao': r.get('descricao'),
                    'total': 0,
                    'estoque_minimo': None,
                }
            qty = r.get('quantidade') or 0
            try:
                qty = int(qty)
            except Exception:
                qty = 0
            produtos_map[key]['total'] += qty
            if possui_colunas_planejamento:
                if produtos_map[key]['estoque_minimo'] is None and r.get('estoque_minimo') is not None:
                    try:
                        produtos_map[key]['estoque_minimo'] = int(r.get('estoque_minimo'))
                    except Exception:
                        produtos_map[key]['estoque_minimo'] = None

        produtos_agg = list(produtos_map.values())
    except Exception as e:
        flash('Erro ao carregar produtos: {}'.format(str(e)), 'error')
        produtos_agg = []
    
    return render_template('produtos.html', produtos=produtos_agg)

@app.route('/produtos/delete/<int:id>', methods=['POST'])
@login_required
def delete_produto(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        # Busca a descrição do produto
        row = g.db.execute(text("SELECT descricao FROM produto WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Produto não encontrado'}), 404
        
        descricao = row['descricao']
        
        # Verifica se há estoque
        rows = g.db.execute(text("""
            SELECT quantidade FROM produto WHERE descricao = :descricao
        """), {"descricao": descricao}).mappings().all()
        
        total = 0
        for r in rows:
            try:
                total += int(r.get('quantidade') or 0)
            except Exception:
                pass
        
        if total > 0:
            return jsonify({'error': 'Não é possível excluir produto com estoque maior que 0 (somatório por locais)'}), 400
        
        # Verifica se existem movimentações vinculadas ao produto
        mov_count = g.db.execute(text("""
            SELECT COUNT(*) as total FROM movimentacao 
            WHERE produto_descricao = :descricao
        """), {"descricao": descricao}).mappings().first()
        
        if mov_count and int(mov_count['total']) > 0:
            return jsonify({
                'error': 'Não é possível excluir este produto pois existem {} movimentação(ões) vinculada(s). Sugerimos inativá-lo em vez de excluir.'.format(int(mov_count['total'])),
                'has_movimentacoes': True,
                'total_movimentacoes': int(mov_count['total'])
            }), 400
        
        # Remove todos os produtos com a mesma descrição
        g.db.execute(text("DELETE FROM produto WHERE descricao = :descricao"), {"descricao": descricao})
        g.db.commit()
        return jsonify({'success': 'Produto excluído com sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/produtos/details/<int:id>', methods=['GET'])
@login_required
def produto_details(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        # Busca a descrição do produto
        row = g.db.execute(text("SELECT descricao FROM produto WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Produto não encontrado'}), 404
        
        descricao = row['descricao']
        
        # Busca todos os detalhes por local
        rows = g.db.execute(text("""
            SELECT id, local, quantidade FROM produto WHERE descricao = :descricao ORDER BY local ASC
        """), {"descricao": descricao}).mappings().all()
        
        detalhes = []
        for r in rows:
            detail = dict(r)
            if 'local' not in detail or detail['local'] is None:
                detail['local'] = ''
            detail['quantidade'] = int(detail.get('quantidade') or 0)
            detalhes.append(detail)
        
        return jsonify({'descricao': descricao, 'detalhes': detalhes})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/produtos/edit/<int:id>', methods=['GET'])
@login_required
def get_produto(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        if _produto_tem_colunas_planejamento():
            row = g.db.execute(text("""
                SELECT id, descricao, quantidade, local, estoque_minimo
                FROM produto WHERE id = :id
            """), {"id": id}).mappings().first()
        else:
            row = g.db.execute(text("""
                SELECT id, descricao, quantidade, local
                FROM produto WHERE id = :id
            """), {"id": id}).mappings().first()
        
        if not row:
            return jsonify({'error': 'Produto não encontrado'}), 404
        
        produto = dict(row)
        if 'local' not in produto or produto['local'] is None:
            produto['local'] = ''
        if 'estoque_minimo' not in produto:
            produto['estoque_minimo'] = None
        
        return jsonify(produto)
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/locais', methods=['GET', 'POST'])
@login_required
def locais():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))
    
    if request.method == 'POST':
        try:
            data = {
                'nome': request.form['nome'],
                'descricao': request.form.get('descricao','')
            }
            
            if 'id' in request.form and request.form['id']:
                g.db.execute(text("""
                    UPDATE local_estoque SET nome = :nome, descricao = :descricao WHERE id = :id
                """), {**data, "id": request.form['id']})
            else:
                g.db.execute(text("""
                    INSERT INTO local_estoque (nome, descricao) VALUES (:nome, :descricao)
                """), data)
            
            g.db.commit()
            flash('Local salvo com sucesso!', 'success')
        except Exception as e:
            flash('Erro ao salvar local: {}'.format(str(e)), 'error')
        
        return redirect(url_for('locais'))
    
    try:
        rows = g.db.execute(text("SELECT * FROM local_estoque ORDER BY id ASC")).mappings().all()
        locais_data = []
        
        for row in rows:
            local = dict(row)
            nome_local = local.get('nome')
            
            # Calcula total de produtos no local
            total = 0
            try:
                prod_rows = g.db.execute(text("""
                    SELECT quantidade FROM produto WHERE local = :local
                """), {"local": nome_local}).mappings().all()
                
                for pr in prod_rows:
                    try:
                        total += int(pr.get('quantidade') or 0)
                    except Exception:
                        pass
            except Exception:
                pass
            
            local['total'] = total
            locais_data.append(local)
            
    except Exception as e:
        flash('Erro ao carregar locais: {}'.format(str(e)), 'error')
        locais_data = []
    
    return render_template('locais.html', locais=locais_data)

@app.route('/locais/edit/<int:id>', methods=['GET'])
@login_required
def get_local(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        row = g.db.execute(text("SELECT * FROM local_estoque WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Local não encontrado'}), 404
        return jsonify(dict(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/locais/delete/<int:id>', methods=['POST'])
@login_required
def delete_local(id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403
    
    try:
        # Busca o nome do local
        row = g.db.execute(text("SELECT nome FROM local_estoque WHERE id = :id"), {"id": id}).mappings().first()
        if not row:
            return jsonify({'error': 'Local não encontrado'}), 404
        
        nome = row['nome']
        
        # Verifica se há produtos no local
        prod_rows = g.db.execute(text("""
            SELECT quantidade FROM produto WHERE local = :local
        """), {"local": nome}).mappings().all()
        
        total = 0
        for pr in prod_rows:
            try:
                total += int(pr.get('quantidade') or 0)
            except Exception:
                pass
        
        if total > 0:
            return jsonify({'error': 'Não é possível excluir local com produtos em estoque (>0).'}), 400
        
        # Verifica se existem movimentações vinculadas ao local
        mov_count = g.db.execute(text("""
            SELECT COUNT(*) as total FROM movimentacao 
            WHERE local = :local
        """), {"local": nome}).mappings().first()
        
        if mov_count and int(mov_count['total']) > 0:
            return jsonify({
                'error': 'Não é possível excluir este local pois existem {} movimentação(ões) vinculada(s). O local só pode ser excluído quando não houver movimentações nem estoque.'.format(int(mov_count['total'])),
                'has_movimentacoes': True,
                'total_movimentacoes': int(mov_count['total'])
            }), 400
        
        g.db.execute(text("DELETE FROM local_estoque WHERE id = :id"), {"id": id})
        g.db.commit()
        return jsonify({'success': 'Local excluído com sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/movimentacoes', methods=['GET', 'POST'])
@login_required
def movimentacoes():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))
    
    if request.method == 'POST':
        try:
            tipo = request.form['tipo']
            produto_id = request.form.get('produto_id')
            quantidade = int(request.form['quantidade'])
            local = request.form['local']
            parceiro_id = request.form.get('parceiro_id') or None

            # Validação do produto_id
            if not produto_id or produto_id == '' or produto_id == '0':
                flash('Por favor, selecione um produto.', 'error')
                return redirect(url_for('movimentacoes'))

            # Busca a descrição do produto
            row = g.db.execute(text("SELECT descricao FROM produto WHERE id = :id"), {"id": produto_id}).mappings().first()
            if not row:
                flash('Produto não encontrado', 'error')
                return redirect(url_for('movimentacoes'))
            
            descricao = row['descricao']

            # Verifica se já existe produto neste local
            existing = g.db.execute(text("""
                SELECT id, quantidade FROM produto WHERE descricao = :descricao AND local = :local
            """), {"descricao": descricao, "local": local}).mappings().first()
            
            if existing:
                # Atualiza quantidade existente
                existing_data = dict(existing)
                new_q = int(existing_data.get('quantidade') or 0) + (quantidade if tipo == 'entrada' else -quantidade)
                
                if new_q < 0:
                    flash('Estoque insuficiente para saída', 'error')
                    return redirect(url_for('movimentacoes'))
                
                g.db.execute(text("""
                    UPDATE produto SET quantidade = :quantidade WHERE id = :id
                """), {"quantidade": new_q, "id": existing_data['id']})
                
                # Verifica se a atualização resultou em estoque negativo
                check = g.db.execute(text("""
                    SELECT quantidade FROM produto WHERE id = :id
                """), {"id": existing_data['id']}).mappings().first()
                
                if check and int(check.get('quantidade') or 0) < 0:
                    g.db.rollback()
                    flash('Operação cancelada: atualização resultou em estoque negativo.', 'error')
                    return redirect(url_for('movimentacoes'))
            else:
                # Cria novo registro de produto no local
                q = quantidade if tipo == 'entrada' else -quantidade
                if q < 0:
                    flash('Estoque insuficiente para saída', 'error')
                    return redirect(url_for('movimentacoes'))
                
                g.db.execute(text("""
                    INSERT INTO produto (descricao, local, quantidade) VALUES (:descricao, :local, :quantidade)
                """), {"descricao": descricao, "local": local, "quantidade": q})

            # Registra a movimentação
            parceiro_tipo = request.form.get('parceiro_tipo') or ('fornecedor' if tipo == 'entrada' else 'cliente')
            mov_data = {
                'tipo': tipo,
                'produto_id': int(produto_id) if produto_id else None,
                'produto_descricao': descricao,
                'quantidade': quantidade,
                'local': local,
                'parceiro_id': parceiro_id,
                'parceiro_tipo': parceiro_tipo
            }
            
            g.db.execute(text("""
                INSERT INTO movimentacao (tipo, produto_id, produto_descricao, quantidade, local, parceiro_id, parceiro_tipo)
                VALUES (:tipo, :produto_id, :produto_descricao, :quantidade, :local, :parceiro_id, :parceiro_tipo)
            """), mov_data)
            
            g.db.commit()
            flash('Movimentação registrada', 'success')
        except Exception as e:
            g.db.rollback()
            flash('Erro ao registrar movimentação: {}'.format(str(e)), 'error')
        
        return redirect(url_for('movimentacoes'))

    # GET - Listar movimentações com filtros
    filter_tipo = request.args.get('filter_tipo') or 'both'
    filter_date_from = request.args.get('date_from') or None
    filter_date_to = request.args.get('date_to') or None
    
    if filter_tipo not in ('both', 'entrada', 'saida'):
        filter_tipo = 'both'
    
    try:
        # Monta query com filtros
        query = "SELECT * FROM movimentacao WHERE 1=1"
        params = {}
        
        if filter_tipo != 'both':
            query += " AND tipo = :tipo"
            params['tipo'] = filter_tipo
        
        if filter_date_from:
            try:
                df = datetime.strptime(filter_date_from, '%Y-%m-%d')
                query += " AND created_at >= :date_from"
                params['date_from'] = df.isoformat()
            except Exception:
                pass
        
        if filter_date_to:
            try:
                dt = datetime.strptime(filter_date_to, '%Y-%m-%d')
                dt = dt.replace(hour=23, minute=59, second=59)
                query += " AND created_at <= :date_to"
                params['date_to'] = dt.isoformat()
            except Exception:
                pass
        
        query += " ORDER BY id DESC"
        
        rows = g.db.execute(text(query), params).mappings().all()
        movimentacoes_data = []
        
        SENTINEL_TRANSFER_PARTNER = 2147483647
        
        for row in rows:
            mov = dict(row)
            
            # Resolve nome do parceiro
            try:
                parceiro_id_raw = mov.get('parceiro_id')
                parceiro_tipo = mov.get('parceiro_tipo')
                
                try:
                    parceiro_id = int(parceiro_id_raw) if parceiro_id_raw is not None else None
                except Exception:
                    parceiro_id = None

                if parceiro_id == SENTINEL_TRANSFER_PARTNER or parceiro_tipo == 'transferencia':
                    mov['parceiro'] = 'TRANSFERENCIA ENTRE ESTOQUES'
                elif parceiro_id and parceiro_tipo == 'fornecedor':
                    f_row = g.db.execute(text("""
                        SELECT nome_fantasia FROM fornecedor WHERE id = :id
                    """), {"id": parceiro_id}).mappings().first()
                    mov['parceiro'] = f_row['nome_fantasia'] if f_row else ''
                elif parceiro_id and parceiro_tipo == 'cliente':
                    c_row = g.db.execute(text("""
                        SELECT nome FROM cliente WHERE id = :id
                    """), {"id": parceiro_id}).mappings().first()
                    mov['parceiro'] = c_row['nome'] if c_row else ''
                else:
                    mov['parceiro'] = mov.get('parceiro') or ''
            except Exception:
                mov['parceiro'] = mov.get('parceiro') or ''
            
            # Formata data
            try:
                ca = mov.get('created_at')
                if ca:
                    if isinstance(ca, str):
                        dt = datetime.fromisoformat(ca)
                    else:
                        dt = ca
                    mov['created_at'] = dt.strftime('%d/%m/%Y %H:%M:%S')
            except Exception:
                pass
            
            movimentacoes_data.append(mov)
        
        # Busca dados para os selects
        locais_rows = g.db.execute(text("SELECT * FROM local_estoque ORDER BY id ASC")).mappings().all()
        locais_data = [dict(row) for row in locais_rows]
        
        # Produtos únicos para o select
        prod_rows = g.db.execute(text("""
            SELECT
                MIN(id) AS id,
                descricao
            FROM produto
            WHERE descricao IS NOT NULL AND descricao != ''
            GROUP BY descricao
            ORDER BY descricao ASC
        """)).mappings().all()

        # Mapa de estoque por descrição/local para exibir no select de locais
        estoque_rows = g.db.execute(text("""
            SELECT
                descricao,
                TRIM(local) AS local,
                SUM(COALESCE(quantidade, 0)) AS quantidade
            FROM produto
            WHERE descricao IS NOT NULL
              AND descricao != ''
              AND COALESCE(TRIM(local), '') != ''
            GROUP BY descricao, TRIM(local)
        """)).mappings().all()

        estoque_por_descricao = {}
        for row in estoque_rows:
            item = dict(row)
            descricao = item.get('descricao')
            local = item.get('local')
            if not descricao or not local:
                continue
            if descricao not in estoque_por_descricao:
                estoque_por_descricao[descricao] = {}
            try:
                estoque_por_descricao[descricao][local] = int(item.get('quantidade') or 0)
            except Exception:
                estoque_por_descricao[descricao][local] = 0
        
        produtos_list = []
        produtos_estoque_map = {}
        for row in prod_rows:
            produto = dict(row)
            if produto.get('descricao'):  # Só adiciona se tem descrição
                # Adiciona sample_id para compatibilidade com o template
                produto['sample_id'] = produto['id']
                produtos_list.append(produto)
            produtos_estoque_map[str(produto['sample_id'])] = estoque_por_descricao.get(produto['descricao'], {})
        
        fornecedores_rows = g.db.execute(text("SELECT id, nome_fantasia FROM fornecedor WHERE ativo = TRUE")).mappings().all()
        fornecedores = [dict(row) for row in fornecedores_rows]
        
        clientes_rows = g.db.execute(text("SELECT id, nome FROM cliente WHERE ativo = TRUE")).mappings().all()
        clientes = [dict(row) for row in clientes_rows]
        
    except Exception as e:
        flash('Erro ao carregar movimentações: {}'.format(str(e)), 'error')
        movimentacoes_data = []
        locais_data = []
        produtos_list = []
        produtos_estoque_map = {}
        fornecedores = []
        clientes = []
    
    return render_template('movimentacoes.html', 
                         movimentacoes=movimentacoes_data, 
                         locais=locais_data, 
                         produtos_list=produtos_list, 
                         produtos_estoque_map=produtos_estoque_map,
                         fornecedores=fornecedores, 
                         clientes=clientes, 
                         filter_tipo=filter_tipo, 
                         filter_date_from=filter_date_from, 
                         filter_date_to=filter_date_to)

@app.route('/movimentacoes/transfer', methods=['POST'])
@login_required
def transfer_movimentacao():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))

    try:
        produto_id = request.form.get('produto_id')
        try:
            quantidade = int(request.form.get('quantidade') or 0)
        except Exception:
            flash('Quantidade inválida.', 'error')
            return redirect(url_for('movimentacoes'))
        
        local_from = request.form.get('local_from')
        local_to = request.form.get('local_to')

        # Validações básicas
        if not produto_id or not local_from or not local_to or quantidade <= 0:
            flash('Preencha todos os campos corretamente.', 'error')
            return redirect(url_for('movimentacoes'))
        
        if local_from == local_to:
            flash('Escolha estoques diferentes (origem e destino).', 'error')
            return redirect(url_for('movimentacoes'))

        # Busca descrição do produto
        prod_row = g.db.execute(text("SELECT descricao FROM produto WHERE id = :id"), {"id": produto_id}).mappings().first()
        if not prod_row:
            flash('Produto não encontrado.', 'error')
            return redirect(url_for('movimentacoes'))
        
        descricao = prod_row['descricao']

        # Verifica estoque no local de origem
        src_row = g.db.execute(text("""
            SELECT id, quantidade FROM produto WHERE descricao = :descricao AND local = :local
        """), {"descricao": descricao, "local": local_from}).mappings().first()
        
        if not src_row:
            flash('Produto não possui estoque no local de origem.', 'error')
            return redirect(url_for('movimentacoes'))
        
        src_data = dict(src_row)
        try:
            src_q = int(src_data.get('quantidade') or 0)
        except Exception:
            src_q = 0
        
        if src_q < quantidade:
            flash('Estoque insuficiente no local de origem.', 'error')
            return redirect(url_for('movimentacoes'))

        # Atualiza estoque no local de origem (subtrai)
        new_src_q = src_q - quantidade
        g.db.execute(text("""
            UPDATE produto SET quantidade = :quantidade WHERE id = :id
        """), {"quantidade": new_src_q, "id": src_data['id']})

        # Verifica se a atualização foi válida
        check_src = g.db.execute(text("""
            SELECT quantidade FROM produto WHERE id = :id
        """), {"id": src_data['id']}).mappings().first()
        
        if check_src and int(check_src.get('quantidade') or 0) < 0:
            g.db.rollback()
            flash('Operação cancelada: transferência deixaria estoque negativo no local de origem.', 'error')
            return redirect(url_for('movimentacoes'))

        # Verifica se já existe produto no local de destino
        dst_row = g.db.execute(text("""
            SELECT id, quantidade FROM produto WHERE descricao = :descricao AND local = :local
        """), {"descricao": descricao, "local": local_to}).mappings().first()
        
        if dst_row:
            # Atualiza estoque existente no destino (soma)
            dst_data = dict(dst_row)
            try:
                dst_q = int(dst_data.get('quantidade') or 0)
            except Exception:
                dst_q = 0
            
            g.db.execute(text("""
                UPDATE produto SET quantidade = :quantidade WHERE id = :id
            """), {"quantidade": dst_q + quantidade, "id": dst_data['id']})
            
            # Verifica se a atualização foi válida
            check_dst = g.db.execute(text("""
                SELECT quantidade FROM produto WHERE id = :id
            """), {"id": dst_data['id']}).mappings().first()
            
            if check_dst and int(check_dst.get('quantidade') or 0) < 0:
                g.db.rollback()
                flash('Operação cancelada: atualização inválida do estoque do destino.', 'error')
                return redirect(url_for('movimentacoes'))
        else:
            # Cria novo registro no local de destino
            g.db.execute(text("""
                INSERT INTO produto (descricao, local, quantidade) VALUES (:descricao, :local, :quantidade)
            """), {"descricao": descricao, "local": local_to, "quantidade": quantidade})
            
            # Verifica se a inserção foi válida
            new_row = g.db.execute(text("""
                SELECT id, quantidade FROM produto WHERE descricao = :descricao AND local = :local
            """), {"descricao": descricao, "local": local_to}).mappings().first()
            
            if new_row and int(new_row.get('quantidade') or 0) < 0:
                g.db.rollback()
                flash('Operação cancelada: inserção inválida no destino.', 'error')
                return redirect(url_for('movimentacoes'))

        # Registra as movimentações (saída do local origem e entrada no destino)
        SENTINEL_TRANSFER_PARTNER = 2147483647
        
        # Movimentação de saída
        g.db.execute(text("""
            INSERT INTO movimentacao (tipo, produto_id, produto_descricao, quantidade, local, parceiro_id, parceiro_tipo)
            VALUES (:tipo, :produto_id, :produto_descricao, :quantidade, :local, :parceiro_id, :parceiro_tipo)
        """), {
            'tipo': 'saida',
            'produto_id': int(produto_id) if produto_id else None,
            'produto_descricao': descricao,
            'quantidade': quantidade,
            'local': local_from,
            'parceiro_id': SENTINEL_TRANSFER_PARTNER,
            'parceiro_tipo': 'transferencia'
        })
        
        # Movimentação de entrada
        g.db.execute(text("""
            INSERT INTO movimentacao (tipo, produto_id, produto_descricao, quantidade, local, parceiro_id, parceiro_tipo)
            VALUES (:tipo, :produto_id, :produto_descricao, :quantidade, :local, :parceiro_id, :parceiro_tipo)
        """), {
            'tipo': 'entrada',
            'produto_id': int(produto_id) if produto_id else None,
            'produto_descricao': descricao,
            'quantidade': quantidade,
            'local': local_to,
            'parceiro_id': SENTINEL_TRANSFER_PARTNER,
            'parceiro_tipo': 'transferencia'
        })
        
        g.db.commit()
        flash('Transferência realizada com sucesso.', 'success')
        
    except Exception as e:
        g.db.rollback()
        flash('Erro ao realizar transferência: {}'.format(str(e)), 'error')
    
    return redirect(url_for('movimentacoes'))

@app.route('/movimentacoes/relatorio', methods=['POST'])
@login_required
def movimentacoes_relatorio():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))

    try:
        report_type = request.form.get('report_type') or 'movimentos'
        date_from = request.form.get('date_from') or None
        date_to = request.form.get('date_to') or None
        out_format = request.form.get('format') or 'excel'

        # Monta query base
        query = "SELECT * FROM movimentacao WHERE 1=1"
        params = {}
        
        if report_type == 'entradas':
            query += " AND tipo = 'entrada'"
        elif report_type == 'saidas':
            query += " AND tipo = 'saida'"
        
        if date_from:
            try:
                df = datetime.strptime(date_from, '%Y-%m-%d')
                query += " AND created_at >= :date_from"
                params['date_from'] = df.isoformat()
            except Exception:
                pass
        
        if date_to:
            try:
                dt = datetime.strptime(date_to, '%Y-%m-%d')
                dt = dt.replace(hour=23, minute=59, second=59)
                query += " AND created_at <= :date_to"
                params['date_to'] = dt.isoformat()
            except Exception:
                pass
        
        query += " ORDER BY id ASC"
        
        rows = g.db.execute(text(query), params).mappings().all()
        movimentacoes_data = []
        
        SENTINEL_TRANSFER_PARTNER = 2147483647
        
        for row in rows:
            mov = dict(row)
            
            # Resolve nome do parceiro
            try:
                parceiro_id_raw = mov.get('parceiro_id')
                parceiro_tipo = mov.get('parceiro_tipo')
                
                try:
                    parceiro_id = int(parceiro_id_raw) if parceiro_id_raw is not None else None
                except Exception:
                    parceiro_id = None

                if parceiro_id == SENTINEL_TRANSFER_PARTNER or parceiro_tipo == 'transferencia':
                    mov['parceiro'] = 'TRANSFERENCIA ENTRE ESTOQUES'
                elif parceiro_id and parceiro_tipo == 'fornecedor':
                    f_row = g.db.execute(text("""
                        SELECT nome_fantasia FROM fornecedor WHERE id = :id
                    """), {"id": parceiro_id}).mappings().first()
                    mov['parceiro'] = f_row['nome_fantasia'] if f_row else ''
                elif parceiro_id and parceiro_tipo == 'cliente':
                    c_row = g.db.execute(text("""
                        SELECT nome FROM cliente WHERE id = :id
                    """), {"id": parceiro_id}).mappings().first()
                    mov['parceiro'] = c_row['nome'] if c_row else ''
                else:
                    mov['parceiro'] = mov.get('parceiro') or ''
            except Exception:
                mov['parceiro'] = mov.get('parceiro') or ''
            
            # Formata data
            try:
                ca = mov.get('created_at')
                if ca:
                    if isinstance(ca, str):
                        dt = datetime.fromisoformat(ca)
                    else:
                        dt = ca
                    mov['created_at'] = dt.strftime('%d/%m/%Y %H:%M:%S')
            except Exception:
                pass
            
            movimentacoes_data.append(mov)

        # Gera CSV
        output = io.StringIO()
        writer = csv.writer(output)

        if report_type in ('por_produto', 'por_local', 'por_cliente', 'por_fornecedor'):
            if report_type == 'por_produto':
                agg = {}
                for m in movimentacoes_data:
                    key = m.get('produto_descricao') or 'N/A'
                    agg.setdefault(key, 0)
                    try:
                        agg[key] += int(m.get('quantidade') or 0)
                    except Exception:
                        pass
                writer.writerow(['Produto','Quantidade Total'])
                for k,v in agg.items():
                    writer.writerow([k, v])
            elif report_type == 'por_local':
                agg = {}
                for m in movimentacoes_data:
                    key = m.get('local') or 'N/A'
                    agg.setdefault(key, 0)
                    try:
                        agg[key] += int(m.get('quantidade') or 0)
                    except Exception:
                        pass
                writer.writerow(['Local','Quantidade Total'])
                for k,v in agg.items():
                    writer.writerow([k, v])
            elif report_type == 'por_cliente':
                agg = {}
                for m in movimentacoes_data:
                    if (m.get('parceiro_tipo') == 'transferencia') or (str(m.get('parceiro') or '').upper() == 'TRANSFERENCIA ENTRE ESTOQUES'):
                        continue
                    if m.get('tipo') != 'saida':
                        continue
                    key = m.get('parceiro') or 'N/A'
                    agg.setdefault(key, 0)
                    try:
                        agg[key] += int(m.get('quantidade') or 0)
                    except Exception:
                        pass
                writer.writerow(['Cliente/Parceiro','Quantidade Total Saídas'])
                for k,v in agg.items():
                    writer.writerow([k, v])
            elif report_type == 'por_fornecedor':
                agg = {}
                for m in movimentacoes_data:
                    if (m.get('parceiro_tipo') == 'transferencia') or (str(m.get('parceiro') or '').upper() == 'TRANSFERENCIA ENTRE ESTOQUES'):
                        continue
                    if m.get('tipo') != 'entrada':
                        continue
                    key = m.get('parceiro') or 'N/A'
                    agg.setdefault(key, 0)
                    try:
                        agg[key] += int(m.get('quantidade') or 0)
                    except Exception:
                        pass
                writer.writerow(['Fornecedor/Parceiro','Quantidade Total Entradas'])
                for k,v in agg.items():
                    writer.writerow([k, v])
        else:
            writer.writerow(['ID','Tipo','Produto','Quantidade','Local','Parceiro','Data'])
            for m in movimentacoes_data:
                writer.writerow([m.get('id'), m.get('tipo'), m.get('produto_descricao'), m.get('quantidade'), m.get('local'), m.get('parceiro'), m.get('created_at')])

        csv_data = output.getvalue()
        output.close()

        # ── EXCEL FORMATADO (openpyxl) ──────────────────────────────
        if out_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = 'Relatório'

                # Estilos
                header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
                header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_font = Font(name='Calibri', size=10)
                cell_align = Alignment(vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin', color='B4C6E7'),
                    right=Side(style='thin', color='B4C6E7'),
                    top=Side(style='thin', color='B4C6E7'),
                    bottom=Side(style='thin', color='B4C6E7')
                )
                zebra_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
                entrada_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                saida_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                entrada_font = Font(name='Calibri', size=10, color='006100')
                saida_font = Font(name='Calibri', size=10, color='9C0006')

                # Título do relatório
                titulos_relatorio = {
                    'movimentos': 'Relatório de Movimentações',
                    'entradas': 'Relatório de Entradas',
                    'saidas': 'Relatório de Saídas',
                    'por_produto': 'Relatório por Produto',
                    'por_local': 'Relatório por Local',
                    'por_cliente': 'Relatório por Cliente',
                    'por_fornecedor': 'Relatório por Fornecedor'
                }
                titulo = titulos_relatorio.get(report_type, 'Relatório de Movimentações')

                # Linha de título
                ws.merge_cells('A1:G1')
                title_cell = ws['A1']
                title_cell.value = titulo
                title_cell.font = Font(name='Calibri', bold=True, size=14, color='2F5496')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')

                # Linha de data
                ws.merge_cells('A2:G2')
                date_cell = ws['A2']
                periodo = ''
                if date_from:
                    periodo += 'De: {} '.format(date_from)
                if date_to:
                    periodo += 'Até: {}'.format(date_to)
                if not periodo:
                    periodo = 'Período: Todos os registros'
                date_cell.value = 'Gerado em: {} | {}'.format(datetime.now().strftime('%d/%m/%Y %H:%M'), periodo)
                date_cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
                date_cell.alignment = Alignment(horizontal='center')

                # Determina cabeçalhos e dados conforme tipo
                if report_type in ('por_produto', 'por_local', 'por_cliente', 'por_fornecedor'):
                    # Relatórios agregados
                    if report_type == 'por_produto':
                        headers = ['Produto', 'Quantidade Total']
                    elif report_type == 'por_local':
                        headers = ['Local', 'Quantidade Total']
                    elif report_type == 'por_cliente':
                        headers = ['Cliente/Parceiro', 'Quantidade Total Saídas']
                    else:
                        headers = ['Fornecedor/Parceiro', 'Quantidade Total Entradas']

                    # Escreve cabeçalho na linha 4
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border

                    # Reconstrói dados agregados
                    agg = {}
                    for m in movimentacoes_data:
                        if report_type == 'por_produto':
                            key = m.get('produto_descricao') or 'N/A'
                        elif report_type == 'por_local':
                            key = m.get('local') or 'N/A'
                        elif report_type == 'por_cliente':
                            if m.get('parceiro_tipo') == 'transferencia' or m.get('tipo') != 'saida':
                                continue
                            key = m.get('parceiro') or 'N/A'
                        else:
                            if m.get('parceiro_tipo') == 'transferencia' or m.get('tipo') != 'entrada':
                                continue
                            key = m.get('parceiro') or 'N/A'
                        agg.setdefault(key, 0)
                        try:
                            agg[key] += int(m.get('quantidade') or 0)
                        except Exception:
                            pass

                    row_num = 5
                    for k, v in agg.items():
                        ws.cell(row=row_num, column=1, value=k).font = cell_font
                        ws.cell(row=row_num, column=1).alignment = cell_align
                        ws.cell(row=row_num, column=1).border = thin_border
                        ws.cell(row=row_num, column=2, value=v).font = Font(name='Calibri', size=10, bold=True)
                        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=row_num, column=2).border = thin_border
                        if (row_num - 5) % 2 == 1:
                            for c in range(1, 3):
                                ws.cell(row=row_num, column=c).fill = zebra_fill
                        row_num += 1

                    # Ajusta largura
                    ws.column_dimensions['A'].width = 40
                    ws.column_dimensions['B'].width = 20

                    # Congela cabeçalho
                    ws.freeze_panes = 'A5'

                else:
                    # Relatório detalhado
                    headers = ['ID', 'Tipo', 'Produto', 'Quantidade', 'Local', 'Parceiro', 'Data']
                    col_widths = [8, 12, 35, 14, 22, 28, 20]

                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_align
                        cell.border = thin_border
                        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

                    row_num = 5
                    for m in movimentacoes_data:
                        values = [
                            m.get('id'),
                            str(m.get('tipo') or '').upper(),
                            m.get('produto_descricao') or '',
                            m.get('quantidade'),
                            m.get('local') or '',
                            m.get('parceiro') or '',
                            m.get('created_at') or ''
                        ]
                        for col_idx, val in enumerate(values, 1):
                            cell = ws.cell(row=row_num, column=col_idx, value=val)
                            cell.font = cell_font
                            cell.alignment = cell_align
                            cell.border = thin_border

                        # Colorir linha conforme tipo
                        tipo = str(m.get('tipo') or '').lower()
                        if tipo == 'entrada':
                            for c in range(1, 8):
                                ws.cell(row=row_num, column=c).fill = entrada_fill
                            ws.cell(row=row_num, column=2).font = entrada_font
                        elif tipo == 'saida':
                            for c in range(1, 8):
                                ws.cell(row=row_num, column=c).fill = saida_fill
                            ws.cell(row=row_num, column=2).font = saida_font

                        # Quantidade centralizada
                        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='center')
                        # ID centralizado
                        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center', vertical='center')

                        row_num += 1

                    # Congela cabeçalho
                    ws.freeze_panes = 'A5'

                    # Linha de totais
                    total_row = row_num
                    ws.cell(row=total_row, column=3, value='TOTAL DE REGISTROS:').font = Font(name='Calibri', bold=True, size=10)
                    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')
                    ws.cell(row=total_row, column=4, value=len(movimentacoes_data)).font = Font(name='Calibri', bold=True, size=10, color='2F5496')
                    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='center')

                # Gera bytes
                excel_buf = io.BytesIO()
                wb.save(excel_buf)
                excel_buf.seek(0)
                excel_bytes = excel_buf.getvalue()
                excel_buf.close()

                filename = 'relatorio_movimentacoes_{}.xlsx'.format(datetime.now().strftime('%Y%m%d%H%M%S'))
                return Response(
                    excel_bytes,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-disposition': 'attachment; filename={}'.format(filename)}
                )
            except Exception as e:
                print('Erro ao gerar Excel: {}'.format(str(e)))
                # Fallback para CSV
                return Response(
                    csv_data,
                    mimetype='text/csv',
                    headers={'Content-disposition': 'attachment; filename=relatorio_movimentacoes_{}.csv'.format(datetime.now().strftime('%Y%m%d%H%M%S'))}
                )

        # ── PDF (ReportLab) ────────────────────────────────────────
        if out_format == 'pdf':
            filename = 'relatorio_movimentacoes_{}.pdf'.format(datetime.now().strftime('%Y%m%d%H%M%S'))

            if REPORTLAB_AVAILABLE:
                try:
                    buf = io.BytesIO()
                    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                            leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
                    styles = getSampleStyleSheet()
                    elems = []

                    # Título
                    titulos_relatorio = {
                        'movimentos': 'Relatório de Movimentações',
                        'entradas': 'Relatório de Entradas',
                        'saidas': 'Relatório de Saídas',
                        'por_produto': 'Relatório por Produto',
                        'por_local': 'Relatório por Local',
                        'por_cliente': 'Relatório por Cliente',
                        'por_fornecedor': 'Relatório por Fornecedor'
                    }
                    titulo = titulos_relatorio.get(report_type, 'Relatório de Movimentações')
                    elems.append(Paragraph(titulo, styles['Title']))

                    # Subtítulo com período
                    periodo = ''
                    if date_from:
                        periodo += 'De: {} '.format(date_from)
                    if date_to:
                        periodo += 'Até: {}'.format(date_to)
                    if periodo:
                        elems.append(Paragraph(periodo, styles['Normal']))
                    elems.append(Paragraph('Gerado em: {}'.format(datetime.now().strftime('%d/%m/%Y %H:%M')), styles['Normal']))
                    elems.append(Spacer(1, 12))

                    if report_type in ('por_produto', 'por_local', 'por_cliente', 'por_fornecedor'):
                        # Tabela agregada
                        if report_type == 'por_produto':
                            headers_pdf = ['Produto', 'Qtd Total']
                        elif report_type == 'por_local':
                            headers_pdf = ['Local', 'Qtd Total']
                        elif report_type == 'por_cliente':
                            headers_pdf = ['Cliente', 'Qtd Saídas']
                        else:
                            headers_pdf = ['Fornecedor', 'Qtd Entradas']

                        agg = {}
                        for m in movimentacoes_data:
                            if report_type == 'por_produto':
                                key = m.get('produto_descricao') or 'N/A'
                            elif report_type == 'por_local':
                                key = m.get('local') or 'N/A'
                            elif report_type == 'por_cliente':
                                if m.get('parceiro_tipo') == 'transferencia' or m.get('tipo') != 'saida':
                                    continue
                                key = m.get('parceiro') or 'N/A'
                            else:
                                if m.get('parceiro_tipo') == 'transferencia' or m.get('tipo') != 'entrada':
                                    continue
                                key = m.get('parceiro') or 'N/A'
                            agg.setdefault(key, 0)
                            try:
                                agg[key] += int(m.get('quantidade') or 0)
                            except Exception:
                                pass

                        data_table = [headers_pdf]
                        for k, v in agg.items():
                            data_table.append([str(k)[:40], str(v)])
                    else:
                        # Tabela detalhada
                        data_table = [['ID', 'Tipo', 'Produto', 'Qtd', 'Local', 'Parceiro', 'Data']]
                        for m in movimentacoes_data:
                            data_table.append([
                                str(m.get('id') or ''),
                                str(m.get('tipo') or '').upper(),
                                str(m.get('produto_descricao') or '')[:30],
                                str(m.get('quantidade') or ''),
                                str(m.get('local') or '')[:20],
                                str(m.get('parceiro') or '')[:25],
                                str(m.get('created_at') or '')
                            ])

                    table = Table(data_table, repeatRows=1)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#D6E4F0')]),
                    ]))
                    elems.append(table)

                    # Total
                    elems.append(Spacer(1, 10))
                    elems.append(Paragraph('Total de registros: {}'.format(len(data_table) - 1), styles['Normal']))

                    doc.build(elems)
                    pdf_bytes = buf.getvalue()
                    buf.close()
                    return Response(pdf_bytes, mimetype='application/pdf',
                                    headers={'Content-disposition': 'attachment; filename={}'.format(filename)})
                except Exception as e:
                    print('ReportLab PDF generation failed: {}'.format(str(e)))
                    return Response(
                        csv_data,
                        mimetype='text/csv',
                        headers={'Content-disposition': 'attachment; filename=relatorio_movimentacoes_{}.csv'.format(datetime.now().strftime('%Y%m%d%H%M%S'))}
                    )

            # Se ReportLab não estiver disponível, retorna CSV
            return Response(
                csv_data,
                mimetype='text/csv',
                headers={'Content-disposition': 'attachment; filename=relatorio_movimentacoes_{}.csv'.format(datetime.now().strftime('%Y%m%d%H%M%S'))}
            )

        # Formato não reconhecido — fallback CSV
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-disposition': 'attachment; filename=relatorio_movimentacoes_{}.csv'.format(datetime.now().strftime('%Y%m%d%H%M%S'))}
        )

    except Exception as e:
        flash('Erro ao gerar relatório: {}'.format(str(e)), 'error')
        return redirect(url_for('movimentacoes'))

@app.route('/movimentacoes/por_parceiro', methods=['GET'])
@login_required
def movimentacoes_por_parceiro():
    if session['user']['role'] == 'admin':
        return jsonify([])
    
    parceiro_tipo = request.args.get('parceiro_tipo')
    parceiro_id = request.args.get('parceiro_id')
    
    if not parceiro_tipo or not parceiro_id:
        return jsonify([])
    
    try:
        # Monta query baseada no tipo de parceiro
        query = """
            SELECT * FROM movimentacao 
            WHERE parceiro_tipo = :parceiro_tipo AND parceiro_id = :parceiro_id
        """
        params = {'parceiro_tipo': parceiro_tipo, 'parceiro_id': int(parceiro_id)}
        
        if parceiro_tipo == 'cliente':
            query += " AND tipo = 'saida'"
        elif parceiro_tipo == 'fornecedor':
            query += " AND tipo = 'entrada'"
        
        query += " ORDER BY id DESC"
        
        rows = g.db.execute(text(query), params).mappings().all()
        movimentacoes = []
        
        for row in rows:
            mov = dict(row)
            mov['quantidade'] = int(mov.get('quantidade') or 0)
            # Formata data para exibição
            try:
                ca = mov.get('created_at')
                if ca:
                    if isinstance(ca, str):
                        dt = datetime.fromisoformat(ca)
                    else:
                        dt = ca
                    mov['created_at_formatted'] = dt.strftime('%d/%m/%Y %H:%M:%S')
            except Exception:
                mov['created_at_formatted'] = str(mov.get('created_at', ''))
            
            movimentacoes.append(mov)
        
        return jsonify(movimentacoes)
    except Exception as e:
        print('Erro ao buscar movimentações por parceiro: {}'.format(str(e)))
        return jsonify([])

@app.route('/movimentacoes/reverter', methods=['POST'])
@login_required
def movimentacoes_reverter():
    if session['user']['role'] == 'admin':
        flash('Ação restrita a usuários normais.', 'error')
        return redirect(url_for('usuarios'))

    try:
        mov_id = request.form.get('movimentacao_id')
        if not mov_id:
            flash('Movimentação não informada.', 'error')
            return redirect(url_for('movimentacoes'))

        # Busca a movimentação original
        mov_row = g.db.execute(text("""
            SELECT * FROM movimentacao WHERE id = :id
        """), {"id": mov_id}).mappings().first()
        
        if not mov_row:
            flash('Movimentação não encontrada.', 'error')
            return redirect(url_for('movimentacoes'))
        
        mov = dict(mov_row)

        # Verifica se é transferência (não pode ser revertida automaticamente)
        if mov.get('parceiro_tipo') == 'transferencia' or mov.get('parceiro_id') == 2147483647:
            flash('Não é possível reverter transferências entre estoques automaticamente.', 'error')
            return redirect(url_for('movimentacoes'))

        original_tipo = mov.get('tipo')
        if original_tipo not in ('entrada', 'saida'):
            flash('Tipo de movimentação inválido para devolução.', 'error')
            return redirect(url_for('movimentacoes'))

        # Define o tipo inverso
        inverse_tipo = 'entrada' if original_tipo == 'saida' else 'saida'
        produto_id = mov.get('produto_id')
        descricao = mov.get('produto_descricao')
        quantidade = int(mov.get('quantidade') or 0)
        local = mov.get('local') or ''
        parceiro_id = mov.get('parceiro_id')
        parceiro_tipo = mov.get('parceiro_tipo')

        # Verifica se existe produto no local
        existing = g.db.execute(text("""
            SELECT id, quantidade FROM produto WHERE descricao = :descricao AND local = :local
        """), {"descricao": descricao, "local": local}).mappings().first()

        if existing:
            existing_data = dict(existing)
            current_q = int(existing_data.get('quantidade') or 0)
            new_q = current_q + (quantidade if inverse_tipo == 'entrada' else -quantidade)
            
            if new_q < 0:
                flash('Estoque insuficiente para realizar a devolução (resultado negativo).', 'error')
                return redirect(url_for('movimentacoes'))
            
            g.db.execute(text("""
                UPDATE produto SET quantidade = :quantidade WHERE id = :id
            """), {"quantidade": new_q, "id": existing_data['id']})
        else:
            if inverse_tipo == 'entrada':
                g.db.execute(text("""
                    INSERT INTO produto (descricao, local, quantidade) VALUES (:descricao, :local, :quantidade)
                """), {"descricao": descricao, "local": local, "quantidade": quantidade})
            else:
                flash('Estoque não possui registro para remover a quantidade solicitada.', 'error')
                return redirect(url_for('movimentacoes'))

        # Registra a movimentação de devolução
        new_mov = {
            'tipo': inverse_tipo,
            'produto_id': int(produto_id) if produto_id else None,
            'produto_descricao': descricao,
            'quantidade': quantidade,
            'local': local,
            'parceiro_id': parceiro_id,
            'parceiro_tipo': parceiro_tipo
        }
        
        g.db.execute(text("""
            INSERT INTO movimentacao (tipo, produto_id, produto_descricao, quantidade, local, parceiro_id, parceiro_tipo)
            VALUES (:tipo, :produto_id, :produto_descricao, :quantidade, :local, :parceiro_id, :parceiro_tipo)
        """), new_mov)
        
        g.db.commit()
        flash('Devolução registrada com sucesso.', 'success')
        
    except Exception as e:
        g.db.rollback()
        flash('Erro ao registrar reversão: {}'.format(str(e)), 'error')

    return redirect(url_for('movimentacoes'))

# ════════════════════════════════════════════════════════════════
# ALTERNATIVA 1: DASHBOARD DE ANÁLISE DE DADOS
# ════════════════════════════════════════════════════════════════

@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    """Dashboard com indicadores KPI e gráficos analíticos"""
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))
    
    try:
        periodo_dias = int(request.args.get('periodo', 30))
    except:
        periodo_dias = 30
    
    try:
        # KPI 1: Total de entradas (últimos N dias)
        total_entradas = g.db.execute(text("""
            SELECT COUNT(*) as qtd FROM movimentacao 
            WHERE tipo = 'entrada' 
            AND created_at >= now() - INTERVAL '{}' day
        """.format(periodo_dias))).mappings().first()
        total_entradas = int(total_entradas['qtd'] or 0) if total_entradas else 0
        
        # KPI 2: Total de saídas (últimos N dias)
        total_saidas = g.db.execute(text("""
            SELECT COUNT(*) as qtd FROM movimentacao 
            WHERE tipo = 'saida' 
            AND created_at >= now() - INTERVAL '{}' day
        """.format(periodo_dias))).mappings().first()
        total_saidas = int(total_saidas['qtd'] or 0) if total_saidas else 0
        
        # KPI 3: Movimentação por período (últimos 12 meses)
        movimentacao_mensal = g.db.execute(text("""
            SELECT 
                DATE_TRUNC('month', created_at)::DATE as mes,
                tipo,
                COUNT(*) as qtd
            FROM movimentacao
            WHERE created_at >= now() - INTERVAL '12' month
            GROUP BY DATE_TRUNC('month', created_at), tipo
            ORDER BY mes DESC
        """)).mappings().all()
        
        movimentacao_chart = {}
        for row in movimentacao_mensal:
            mes = row['mes'].strftime('%b/%Y') if row['mes'] else 'N/A'
            tipo = row['tipo']
            qtd = int(row['qtd'] or 0)
            if mes not in movimentacao_chart:
                movimentacao_chart[mes] = {'entrada': 0, 'saida': 0}
            movimentacao_chart[mes][tipo] = qtd
        
        # KPI 4: Top 10 produtos com maior saída (últimos N dias)
        top_produtos_saida = g.db.execute(text("""
            SELECT 
                produto_descricao,
                SUM(quantidade) as total_saida
            FROM movimentacao
            WHERE tipo = 'saida' 
            AND parceiro_tipo != 'transferencia'
            AND created_at >= now() - INTERVAL '{}' day
            GROUP BY produto_descricao
            ORDER BY total_saida DESC
            LIMIT 10
        """.format(periodo_dias))).mappings().all()
        top_produtos_saida = [dict(row) for row in top_produtos_saida]
        
        # KPI 5: Top 10 fornecedores por volume de entrada (últimos N dias)
        top_fornecedores = g.db.execute(text("""
            SELECT 
                COALESCE(f.nome_fantasia, 'Desconhecido') as fornecedor,
                SUM(m.quantidade) as total_entrada
            FROM movimentacao m
            LEFT JOIN fornecedor f ON m.parceiro_id = f.id
            WHERE m.tipo = 'entrada' 
            AND m.parceiro_tipo = 'fornecedor'
            AND m.created_at >= now() - INTERVAL '{}' day
            GROUP BY f.nome_fantasia
            ORDER BY total_entrada DESC
            LIMIT 10
        """.format(periodo_dias))).mappings().all()
        top_fornecedores = [dict(row) for row in top_fornecedores]
        
        # KPI 6: Top 10 clientes por volume de saída (últimos N dias)
        top_clientes = g.db.execute(text("""
            SELECT 
                COALESCE(c.nome, 'Desconhecido') as cliente,
                SUM(m.quantidade) as total_saida
            FROM movimentacao m
            LEFT JOIN cliente c ON m.parceiro_id = c.id
            WHERE m.tipo = 'saida' 
            AND m.parceiro_tipo = 'cliente'
            AND m.created_at >= now() - INTERVAL '{}' day
            GROUP BY c.nome
            ORDER BY total_saida DESC
            LIMIT 10
        """.format(periodo_dias))).mappings().all()
        top_clientes = [dict(row) for row in top_clientes]
        
        # KPI 7: Movimentação por local (últimos N dias)
        movimentacao_por_local = g.db.execute(text("""
            SELECT 
                local,
                COUNT(*) as total
            FROM movimentacao
            WHERE created_at >= now() - INTERVAL '{}' day
            GROUP BY local
            ORDER BY total DESC
        """.format(periodo_dias))).mappings().all()
        movimentacao_por_local = [dict(row) for row in movimentacao_por_local]
        
        # KPI 8: Saldo atual por produto
        saldo_produtos = g.db.execute(text("""
            SELECT
                p.descricao,
                SUM(p.quantidade) as saldo_total,
                COUNT(DISTINCT CASE
                    WHEN COALESCE(TRIM(p.local), '') != '' AND COALESCE(p.quantidade, 0) > 0
                    THEN TRIM(p.local)
                    ELSE NULL
                END) as locais,
                COALESCE(m.entrada_qtd, 0) AS entrada_qtd,
                COALESCE(m.saida_qtd, 0) AS saida_qtd,
                CASE
                    WHEN COALESCE(m.saida_qtd, 0) > 0
                    THEN ROUND((COALESCE(m.entrada_qtd, 0)::numeric / COALESCE(m.saida_qtd, 0)::numeric), 2)
                    ELSE NULL
                END AS relacao_entrada_saida
            FROM produto p
            LEFT JOIN (
                SELECT
                    produto_descricao,
                    COALESCE(SUM(CASE WHEN tipo = 'entrada' THEN quantidade ELSE 0 END), 0) AS entrada_qtd,
                    COALESCE(SUM(CASE WHEN tipo = 'saida' THEN quantidade ELSE 0 END), 0) AS saida_qtd
                FROM movimentacao
                WHERE parceiro_tipo != 'transferencia'
                  AND created_at >= now() - INTERVAL '{}' day
                GROUP BY produto_descricao
            ) m ON m.produto_descricao = p.descricao
            WHERE p.descricao IS NOT NULL AND p.descricao != ''
            GROUP BY p.descricao, m.entrada_qtd, m.saida_qtd
            ORDER BY saldo_total DESC
            LIMIT 20
        """.format(periodo_dias))).mappings().all()
        saldo_produtos = [dict(row) for row in saldo_produtos]
        
        dashboard_data = {
            'total_entradas': total_entradas,
            'total_saidas': total_saidas,
            'movimentacao_mensal': movimentacao_chart,
            'top_produtos_saida': top_produtos_saida,
            'top_fornecedores': top_fornecedores,
            'top_clientes': top_clientes,
            'movimentacao_local': movimentacao_por_local,
            'saldo_produtos': saldo_produtos,
            'periodo': periodo_dias
        }
        
    except Exception as e:
        flash('Erro ao carregar dashboard: {}'.format(str(e)), 'error')
        dashboard_data = {
            'total_entradas': 0,
            'total_saidas': 0,
            'movimentacao_mensal': {},
            'top_produtos_saida': [],
            'top_fornecedores': [],
            'top_clientes': [],
            'movimentacao_local': [],
            'saldo_produtos': [],
            'periodo': periodo_dias
        }
    
    return render_template('dashboard.html', data=dashboard_data)


def _produto_tem_colunas_planejamento():
    try:
        row = g.db.execute(text("""
            SELECT COUNT(*) AS total
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = 'produto'
              AND column_name = 'estoque_minimo'
        """)).mappings().first()
        return bool(row and int(row.get('total') or 0) >= 1)
    except Exception:
        return False


def _buscar_base_reposicao(periodo_dias, possui_colunas_planejamento):
    if possui_colunas_planejamento:
        query = text("""
            SELECT
                p.descricao,
                p.saldo_atual,
                p.estoque_minimo,
                COALESCE(m.total_saida_periodo, 0) AS total_saida_periodo
            FROM (
                SELECT
                    descricao,
                    SUM(COALESCE(quantidade, 0)) AS saldo_atual,
                    MAX(COALESCE(estoque_minimo, 0)) AS estoque_minimo
                FROM produto
                WHERE descricao IS NOT NULL AND descricao != ''
                GROUP BY descricao
            ) p
            LEFT JOIN (
                SELECT
                    produto_descricao,
                    COALESCE(SUM(quantidade), 0) AS total_saida_periodo
                FROM movimentacao
                WHERE tipo = 'saida'
                  AND parceiro_tipo != 'transferencia'
                  AND created_at >= now() - make_interval(days => :periodo)
                GROUP BY produto_descricao
            ) m ON m.produto_descricao = p.descricao
            ORDER BY p.descricao ASC
        """)
    else:
        query = text("""
            SELECT
                p.descricao,
                p.saldo_atual,
                0 AS estoque_minimo,
                COALESCE(m.total_saida_periodo, 0) AS total_saida_periodo
            FROM (
                SELECT
                    descricao,
                    SUM(COALESCE(quantidade, 0)) AS saldo_atual
                FROM produto
                WHERE descricao IS NOT NULL AND descricao != ''
                GROUP BY descricao
            ) p
            LEFT JOIN (
                SELECT
                    produto_descricao,
                    COALESCE(SUM(quantidade), 0) AS total_saida_periodo
                FROM movimentacao
                WHERE tipo = 'saida'
                  AND parceiro_tipo != 'transferencia'
                  AND created_at >= now() - make_interval(days => :periodo)
                GROUP BY produto_descricao
            ) m ON m.produto_descricao = p.descricao
            ORDER BY p.descricao ASC
        """)

    rows = g.db.execute(query, {'periodo': periodo_dias}).mappings().all()
    return [dict(row) for row in rows]


# ════════════════════════════════════════════════════════════════
# ALTERNATIVA 6: ANÁLISE PREDITIVA (REPOSIÇÃO)
# ════════════════════════════════════════════════════════════════

@app.route('/dashboard/reposicoes', methods=['GET'])
@login_required
def dashboard_reposicoes():
    if session['user']['role'] == 'admin':
        flash('Acesso restrito a usuários normais.', 'error')
        return redirect(url_for('usuarios'))

    try:
        periodo_dias = int(request.args.get('periodo', 60))
    except Exception:
        periodo_dias = 60

    if periodo_dias not in (30, 60, 90, 180):
        periodo_dias = 60

    try:
        possui_colunas_planejamento = _produto_tem_colunas_planejamento()
        base = _buscar_base_reposicao(periodo_dias, possui_colunas_planejamento)

        itens = []
        for item in base:
            metricas = calcular_metricas_reposicao(
                saldo_atual=item.get('saldo_atual'),
                total_saida_periodo=item.get('total_saida_periodo'),
                dias_periodo=periodo_dias,
                estoque_minimo=item.get('estoque_minimo'),
            )

            itens.append({
                'descricao': item.get('descricao') or 'Sem descrição',
                **metricas
            })

        itens_ordenados = sorted(
            itens,
            key=lambda entry: (
                0 if entry['criticidade'] == 'critico' else (1 if entry['criticidade'] == 'atencao' else 2),
                -entry['sugestao_compra'],
                entry['descricao'].lower(),
            )
        )

        resumo = {
            'total': len(itens_ordenados),
            'criticos': sum(1 for item in itens_ordenados if item['criticidade'] == 'critico'),
            'atencao': sum(1 for item in itens_ordenados if item['criticidade'] == 'atencao'),
            'ok': sum(1 for item in itens_ordenados if item['criticidade'] == 'ok'),
        }

        return render_template(
            'dashboard_reposicoes.html',
            itens=itens_ordenados,
            resumo=resumo,
            periodo=periodo_dias,
            possui_colunas_planejamento=possui_colunas_planejamento,
        )
    except Exception as e:
        flash('Erro ao gerar análise preditiva: {}'.format(str(e)), 'error')
        return render_template(
            'dashboard_reposicoes.html',
            itens=[],
            resumo={'total': 0, 'criticos': 0, 'atencao': 0, 'ok': 0},
            periodo=periodo_dias,
            possui_colunas_planejamento=False,
        )


@app.route('/api/predicao/<int:produto_id>', methods=['GET'])
@login_required
def api_predicao_produto(produto_id):
    if session['user']['role'] == 'admin':
        return jsonify({'error': 'Acesso restrito a usuários normais.'}), 403

    try:
        try:
            periodo_dias = int(request.args.get('periodo', 60))
        except Exception:
            periodo_dias = 60

        row = g.db.execute(
            text("SELECT descricao FROM produto WHERE id = :id"),
            {'id': produto_id}
        ).mappings().first()

        if not row:
            return jsonify({'error': 'Produto não encontrado'}), 404

        descricao = row.get('descricao')
        possui_colunas_planejamento = _produto_tem_colunas_planejamento()
        base = _buscar_base_reposicao(periodo_dias, possui_colunas_planejamento)
        item = next((entry for entry in base if entry.get('descricao') == descricao), None)

        if not item:
            return jsonify({'error': 'Sem dados para previsão'}), 404

        metricas = calcular_metricas_reposicao(
            saldo_atual=item.get('saldo_atual'),
            total_saida_periodo=item.get('total_saida_periodo'),
            dias_periodo=periodo_dias,
            estoque_minimo=item.get('estoque_minimo'),
        )

        return jsonify({
            'produto_id': produto_id,
            'descricao': descricao,
            'periodo_dias': periodo_dias,
            'usa_parametros_cadastrados': possui_colunas_planejamento,
            **metricas
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

if __name__ == '__main__':
    app.run(debug=True)